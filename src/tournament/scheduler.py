import pandas as pd
from datetime import datetime, timedelta
from src.tournament.tournament_manager import TournamentManager
from src.tournament.config import SUBMITTED_TEAMS, TOURNAMENT_DIR, MATCH_DURATIONS, POOL_DISPLAY_NAMES
from typing import List, Dict, Optional
from src.utils.const import Soldier
import plotly.graph_objects as go
import numpy as np
from pathlib import Path

MATCH_DURATIONS = {
    "random_vs_random": 300 + 90,  # ~6.5 minutes
    "ai_vs_ai": 120 + 60,          # ~2.5 minutes
    "random_vs_ai": 120 + 60,      # Using same duration as ai_vs_ai
    "ai_vs_random": 120 + 60,      # Same as random_vs_ai
    "forfeit": 30                 # 30 seconds
}

class MatchScheduler:
    def __init__(self, pool: str, phase: str = None):
        self.tournament = TournamentManager(None, current_pool=pool)  # Just to get matches
        matches = self.tournament.matches
        self.phase = phase
        
        # Filter matches based on current phase if it's set
        if phase:
            if phase == "ALLER":
                self.matches = matches[:len(matches)//2]
            else:  # RETOUR
                self.matches = matches[len(matches)//2:]
        else:
            self.matches = matches
            
        self.pool = pool
        self.PHASE_BREAK_DURATION = 600  # 15 minutes
        self.MIN_BREAK_DURATION = 300     # 5 minutes
        self.output_dir = TOURNAMENT_DIR / "schedules" / f"pool_{pool}"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _classify_match(self, team1: str, team2: str, forfeit: Optional[Soldier]) -> str:
        """Determine match type based on teams and forfeit status."""
        if forfeit:
            return "forfeit"
        
        team1_is_ai = team1 in SUBMITTED_TEAMS
        team2_is_ai = team2 in SUBMITTED_TEAMS

        if team1_is_ai and team2_is_ai:
            return "ai_vs_ai"
        elif team1_is_ai:
            return "ai_vs_random"
        elif team2_is_ai:
            return "random_vs_ai"
        else:
            return "random_vs_random"

    def generate_schedule(self, start_time: datetime) -> List[Dict]:
        """Generate a schedule for all matches starting at given time."""
        current_time = start_time
        schedule = []

        for i, (team1, team2, forfeit) in enumerate(self.matches, 1):
            match_type = self._classify_match(team1, team2, forfeit)
            duration = MATCH_DURATIONS[match_type]
            
            match_info = {
                "round": i,
                "start_time": current_time,
                "end_time": current_time + timedelta(seconds=duration),
                "team1": team1,
                "team2": team2,
                "match_type": match_type,
                "duration": duration,
                "phase": self.phase
            }
            
            schedule.append(match_info)
            current_time = match_info["end_time"]

        return self._add_breaks(schedule)

    def _add_breaks(self, schedule: List[Dict]) -> List[Dict]:
        """Add breaks between phases."""
        enhanced_schedule = []
        # last_match_end = None

        for match in schedule:
            current_time = match["start_time"]
            
            # # Add break between phases only
            # if enhanced_schedule and match["phase"] != enhanced_schedule[-1]["phase"]:
            #     phase_break = {
            #         "round": "PAUSE",
            #         "start_time": current_time,
            #         "end_time": current_time + timedelta(seconds=self.PHASE_BREAK_DURATION),
            #         "team1": "PAUSE ENTRE PHASES",
            #         "team2": "",
            #         "match_type": "break",
            #         "duration": self.PHASE_BREAK_DURATION,
            #         "phase": "PAUSE"
            #     }
            #     enhanced_schedule.append(phase_break)
            #     current_time = phase_break["end_time"]

            # Update match timing
            match["start_time"] = current_time
            match["end_time"] = current_time + timedelta(seconds=match["duration"])
            enhanced_schedule.append(match)

        return enhanced_schedule

    def format_schedule(self, schedule: List[Dict]) -> str:
        """Format the schedule into a readable string."""
        header = f"""
╔══════════════════════════════════════════════════════════════════════════════
║ PLANNING DES MATCHS - POOL {self.pool}
╠══════════════════════════════════════════════════════════════════════════════
║ Format: [Heure] Team1 vs Team2 (Durée)
╟──────────────────────────────────────────────────────────────────────────────\n"""

        current_phase = None
        formatted = [header]

        for match in schedule:
            if match["phase"] != current_phase:
                current_phase = match["phase"]
                formatted.append(f"\n║ {' PHASE ' + current_phase + ' ':=^74}║\n")

            start = match["start_time"].strftime("%H:%M:%S")
            end = match["end_time"].strftime("%H:%M:%S")
            duration = f"{match['duration']//60:.0f}min {match['duration']%60:.0f}s"

            match_line = (f"║ [{start}-{end}] "
                         f"{match['team1']:20} vs {match['team2']:20} "
                         f"({duration:8})")
            
            formatted.append(match_line + " ║\n")

        formatted.append("╚══════════════════════════════════════════════════════════════════════════════")
        return "".join(formatted)

    def export_to_excel(self, schedule: List[Dict], filename: str):
        """Export schedule to Excel with conditional formatting and styling."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Get pool display name
        pool_name = POOL_DISPLAY_NAMES.get(self.pool, self.pool)
        
        # Color schemes based on gemstones
        GEMSTONE_COLORS = {
            'TOPAZE': {
                'header': '996515',  # Golden brown
                'ai_vs_ai': 'FFE5B4',  # Light topaz
                'random_vs_ai': 'FFD700',  # Darker golden
                'background': 'FFF8DC',  # Very light cream
            },
            'AMETHYSTE': {
                'header': '9966CC',  # Rich purple
                'ai_vs_ai': 'E6E6FA',  # Light amethyst
                'random_vs_ai': 'D8BFD8',  # Darker purple
                'background': 'F8F4FF',  # Very light purple
            }
        }

        # Get color scheme based on pool name
        colors = GEMSTONE_COLORS.get(pool_name, {
            'header': '4472C4',  # Default blue
            'ai_vs_ai': 'BDD7EE',
            'random_vs_ai': 'E2EFDA',
            'background': 'FFFFFF'
        })

        # Column names with proper capitalization
        column_names = {
            'round': 'Round',
            'start_time': 'Début',
            'end_time': 'Fin',
            'team1': 'Équipe 1',
            'team2': 'Équipe 2',
            'duration': 'Durée',
            'phase': 'Phase'
        }

        # Prepare Excel data
        schedule_for_excel = []
        match_types = []
        for match in schedule:
            match_dict = {
                column_names['round']: match['round'],
                column_names['start_time']: match['start_time'],
                column_names['end_time']: match['end_time'],
                column_names['team1']: match['team1'],
                column_names['team2']: match['team2'],
                column_names['duration']: f"{match['duration']//60:.0f}min {match['duration']%60:.0f}s",
                column_names['phase']: match['phase']
            }
            schedule_for_excel.append(match_dict)
            match_types.append(match['match_type'])

        # Create DataFrame
        df = pd.DataFrame(schedule_for_excel)

        # Define styles with gemstone colors
        header_style = Font(bold=True, color='FFFFFF', name='Arial')
        header_fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Match type styles using gemstone colors
        match_styles = {
            'ai_vs_ai': (
                PatternFill(start_color=colors['ai_vs_ai'], end_color=colors['ai_vs_ai'], fill_type='solid'),
                Font(name='Arial')
            ),
            'random_vs_ai': (
                PatternFill(start_color=colors['random_vs_ai'], end_color=colors['random_vs_ai'], fill_type='solid'),
                Font(name='Arial')
            ),
            'ai_vs_random': (  # Same as random_vs_ai
                PatternFill(start_color=colors['random_vs_ai'], end_color=colors['random_vs_ai'], fill_type='solid'),
                Font(name='Arial')
            )
        }

        # Export to Excel with styling
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Planning')
            worksheet = writer.sheets['Planning']
            
            # Set background color for entire sheet
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.fill = PatternFill(start_color=colors['background'], end_color=colors['background'], fill_type='solid')

            # Apply header styles
            for cell in worksheet[1]:
                cell.font = header_style
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Apply match type styles
            for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                match_type = match_types[idx-2]
                if match_type in match_styles:
                    fill, font = match_styles[match_type]
                    for cell in row:
                        cell.fill = fill
                        cell.font = font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left')

            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except:
                        pass
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

def generate_pool_gantt(schedules: dict, pool: str, start_hour: float = 21.0):
    """Generate an interactive timetable-style visualization for pool matches."""
    teams = set()
    
    # Get pool display name and colors
    pool_name = POOL_DISPLAY_NAMES.get(pool, pool)
    
    # Gemstone color schemes
    GEMSTONE_COLORS = {
        'TOPAZE': {
            'primary': 'rgba(255, 215, 0, 0.8)',      # Golden
            'secondary': 'rgba(218, 165, 32, 0.8)',   # Darker golden
            'background': 'rgba(255, 223, 0, 0.1)',   # Very light golden
            'separator': 'rgba(184, 134, 11, 0.8)',   # Golden brown
        },
        'AMETHYSTE': {
            'primary': 'rgba(147, 112, 219, 0.8)',    # Purple
            'secondary': 'rgba(138, 43, 226, 0.8)',   # Darker purple
            'background': 'rgba(230, 230, 250, 0.1)', # Very light purple
            'separator': 'rgba(148, 0, 211, 0.8)',    # Deep purple
        }
    }

    # Get color scheme based on pool name or use default
    colors = GEMSTONE_COLORS.get(pool_name, {
        'primary': 'rgba(103, 58, 183, 0.8)',
        'secondary': 'rgba(0, 150, 136, 0.8)',
        'background': 'rgba(103, 58, 183, 0.1)',
        'separator': 'rgba(244, 67, 54, 0.8)'
    })

    # Create base time from start_hour
    base_time = datetime.now().replace(
        hour=int(start_hour),
        minute=int((start_hour % 1) * 60),
        second=0,
        microsecond=0
    )
    end_time = base_time + timedelta(hours=2)
    
    # Collect teams and sort them
    for phase_schedule in schedules.values():
        for match in phase_schedule:
            teams.add(match['team1'])
            teams.add(match['team2'])
    
    # Create figure with better styling
    fig = go.Figure()
    
    # Enhanced color scheme using gemstone colors
    phase_colors = {
        'ALLER': colors['primary'],
        'RETOUR': colors['secondary']
    }
    
    # Add shaded regions for phases
    phase_change_time = min(m['start_time'] for m in schedules['RETOUR'])
    fig.add_shape(  # ALLER phase background
        type="rect",
        x0=base_time,
        x1=phase_change_time,
        y0=-1,
        y1=len(teams),
        fillcolor=colors['background'],
        line_width=0,
        layer="below"
    )
    fig.add_shape(  # RETOUR phase background
        type="rect",
        x0=phase_change_time,
        x1=end_time,
        y0=-1,
        y1=len(teams),
        fillcolor=colors['background'].replace('0.1', '0.15'),  # Slightly darker
        line_width=0,
        layer="below"
    )
    
    # Add phase separator with enhanced styling
    fig.add_shape(
        type="line",
        x0=phase_change_time,
        x1=phase_change_time,
        y0=-0.5,
        y1=len(teams) - 0.5,
        line=dict(
            color=colors['separator'],
            width=2,
            dash="dash",
        )
    )
    
    # Add phase labels with better positioning and styling
    for phase, (start, end) in {
        'PHASE ALLER': (base_time, phase_change_time),
        'PHASE RETOUR': (phase_change_time, end_time)
    }.items():
        # Calculate center point between two datetimes
        center_time = start + (end - start) / 2
        
        fig.add_annotation(
            x=center_time,  # Use calculated center time
            y=len(teams) + 0.5,
            text=phase,
            showarrow=False,
            font=dict(size=14, color="rgba(0,0,0,0.7)"),
            bgcolor="rgba(255,255,255,0.8)",
            bordercolor="rgba(0,0,0,0.2)",
            borderwidth=1,
            borderpad=4,
            yshift=20
        )

    # Add matches with enhanced styling
    for i, team in enumerate(sorted(teams)):
        for phase, schedule in schedules.items():
            team_matches = [m for m in schedule if team in [m['team1'], m['team2']]]
            
            for match in team_matches:
                opponent = match['team2'] if team == match['team1'] else match['team1']
                # Enhanced hover text
                hover_text = (
                    f"<b>{team}</b> vs <b>{opponent}</b><br>"
                    f"<i>{match['start_time'].strftime('%H:%M')}</i><br>"
                    f"Phase {phase}"
                )
                
                fig.add_trace(go.Scatter(
                    x=[match['start_time']],
                    y=[i],
                    name=team,
                    mode='markers',
                    marker=dict(
                        symbol='square',
                        size=16,
                        color=phase_colors[phase],
                        line=dict(color='white', width=1)
                    ),
                    text=hover_text,
                    hoverinfo='text',
                    hoverlabel=dict(
                        bgcolor='white',
                        font=dict(color='black'),
                        bordercolor=phase_colors[phase]
                    ),
                    showlegend=False
                ))
                
                # Add connection lines with gradient effect
                if len(team_matches) > 1:
                    match_idx = team_matches.index(match)
                    if match_idx < len(team_matches) - 1:
                        next_match = team_matches[match_idx + 1]
                        fig.add_trace(go.Scatter(
                            x=[match['start_time'], next_match['start_time']],
                            y=[i, i],
                            mode='lines',
                            line=dict(
                                color='rgba(0,0,0,0.2)',
                                width=1,
                                dash='dot'
                            ),
                            showlegend=False
                        ))

    # Enhanced layout with adjusted dimensions
    # Get display name for pool
    pool_name = POOL_DISPLAY_NAMES.get(pool, pool)
    
    # Enhanced layout with pool display name
    fig.update_layout(
        title=dict(
            text=f'Planning des matchs - Pool {pool_name}',
            font=dict(size=24, color='rgba(0,0,0,0.8)'),
            x=0.5,
            y=0.95
        ),
        xaxis=dict(
            title='Heure',
            title_font=dict(size=14),
            type='date',
            tickformat='%H:%M',
            dtick=15 * 60 * 1000,
            range=[base_time, end_time],
            showgrid=True,
            gridcolor='rgba(0,0,0,0.1)',
            zeroline=False
        ),
        yaxis=dict(
            title='Équipes',
            title_font=dict(size=14),
            ticktext=sorted(teams),
            tickvals=list(range(len(teams))),
            showgrid=True,
            gridcolor='rgba(0,0,0,0.1)',
            zeroline=False
        ),
        plot_bgcolor='white',
        paper_bgcolor='white',
        height=max(650, len(teams) * 45),  # Slightly increased height per team
        margin=dict(l=100, r=40, t=80, b=40),  # Adjusted margins
        showlegend=True,
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="right",
            x=0.99,
            bgcolor='rgba(255,255,255,0.9)',
            bordercolor='rgba(0,0,0,0.2)',
            borderwidth=1
        ),
        hovermode='closest',
        modebar=dict(
            remove=[
                'select2d', 'lasso2d', 'zoomIn2d', 'zoomOut2d', 'autoScale2d',
                'toggleSpikelines', 'hoverClosestCartesian',
                'hoverCompareCartesian'
            ],
            orientation='v',  # Vertical orientation for better mobile view
            bgcolor='rgba(255,255,255,0.7)',
            color='rgba(0,0,0,0.5)',
        ),
        dragmode='pan',  # Make panning the default instead of zoom box
        hoverdistance=100,  # Increase hover sensitivity
        width=1000,  # Updated from 900 to 1000
        autosize=True,  # Allow responsive scaling above minimum width
    )

    # Save the figure with mobile-friendly configuration
    output_dir = Path(TOURNAMENT_DIR) / "schedules" / f"pool_{pool}"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"gantt_pool_{pool}.html"
    print(f"Saving visualization to: {output_path}")

    # First save the basic HTML with improved mobile config
    fig.write_html(
        str(output_path),
        include_plotlyjs=True,
        full_html=True,
        include_mathjax=False,
        config={
            'responsive': True,
            'scrollZoom': True,
            'displayModeBar': 'hover',
            'modeBarButtonsToRemove': [
                'select2d', 'lasso2d', 'autoScale2d',
                'toggleSpikelines'
            ],
            'displaylogo': False,
            'doubleClick': 'reset+autosize',  # Better double-click behavior
            'toImageButtonOptions': {
                'format': 'png',
                'filename': f'pool_{pool}_schedule',
                'height': 1200,
                'width': 900,
                'scale': 2
            }
        }
    )

    # Then modify the HTML to add our custom styling
    with open(output_path, 'r', encoding='utf-8') as f:
        html_content = f.read()

    # Insert our custom CSS after the <head> tag
    custom_css = """
    <style>
        html, body {
            margin: 0;
            padding: 0;
            width: 100%;
            height: 100%;
            overflow-x: auto;
        }
        .plot-container {
            min-width: 1000px !important;  /* Updated from 900px */
            height: 100vh;
        }
        .js-plotly-plot, .plotly-graph-div {
            height: 100% !important;
        }
        /* Custom zoom controls for mobile */
        @media (hover: none) and (pointer: coarse) {
            .js-plotly-plot .plotly .modebar {
                transform: scale(1.5);
                transform-origin: right top;
            }
        }
    </style>
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=2.0">
    """
    
    html_content = html_content.replace('</head>', f'{custom_css}</head>')

    # Save the modified HTML
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

def create_schedule(pool: str, start_hour: float = 13, phase: str = None):
    """Create and display a schedule for a pool starting at given hour."""
    hours = int(start_hour)
    minutes = int((start_hour % 1) * 60)
    
    start_time = datetime.now().replace(
        hour=hours, minute=minutes, second=0, microsecond=0)
    
    # Create scheduler with phase
    scheduler = MatchScheduler(pool, phase)
    schedule = scheduler.generate_schedule(start_time)
    formatted_schedule = scheduler.format_schedule(schedule)
    
    # Save schedule to files in the pool directory
    phase_suffix = f"_{phase.lower()}" if phase else ""
    base_filename = scheduler.output_dir / f"schedule{phase_suffix}"
    
    with open(f"{base_filename}.txt", "w", encoding="utf-8") as f:
        f.write(formatted_schedule)
    
    print(f"\nSchedule saved to {base_filename}.txt")
    # print(formatted_schedule)
    
    scheduler.export_to_excel(schedule, f"{base_filename}.xlsx")
    
    return schedule

if __name__ == "__main__":
    create_schedule("A", 13, "ALLER")  # Example: Create schedule for pool A, phase ALLER
